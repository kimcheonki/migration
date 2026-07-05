"""
Excel 레포트 생성기
"""
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from src.model.result import FileMigrationResult, MigrationSummary


class ExcelReportGenerator:
    """Excel 레포트 생성기"""
    
    def __init__(self):
        self.wb = None
    
    def generate(self, summary: MigrationSummary, results: List[FileMigrationResult]) -> Workbook:
        """Excel 레포트 생성"""
        self.wb = Workbook()
        
        # 요약 시트 생성
        self._create_summary_sheet(summary)
        
        # 상세 시트 생성
        self._create_detail_sheet(results)
        
        # 이슈 시트 생성
        self._create_issues_sheet(results)
        
        return self.wb
    
    def _create_summary_sheet(self, summary: MigrationSummary):
        """요약 시트 생성"""
        ws = self.wb.active
        ws.title = "요약"
        
        # 헤더 스타일
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 제목
        ws['A1'] = "WEB/WAS 마이그레이션 레포트"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        ws['A2'] = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:B2')
        
        # 서버 정보
        row = 4
        headers = ["항목", "값"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row += 1
        data = [
            ("서버 이름", summary.server_name),
            ("서버 타입", summary.server_type.value),
            ("FROM 버전", summary.from_version),
            ("TO 버전", summary.to_version),
            ("총 파일 수", summary.total_files),
            ("성공 파일", summary.successful_files),
            ("경고 파일", summary.warning_files),
            ("실패 파일", summary.error_files),
            ("수동 검토 파일", summary.manual_review_files),
            ("총 변경사항", summary.total_changes),
            ("총 이슈", summary.total_issues)
        ]
        
        for key, value in data:
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_detail_sheet(self, results: List[FileMigrationResult]):
        """상세 시트 생성"""
        ws = self.wb.create_sheet("상세")
        
        # 헤더
        headers = ["파일 경로", "파일 타입", "상태", "적용 가능", "변경사항 수", "이슈 수"]
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # 데이터
        row = 2
        for result in results:
            ws.cell(row=row, column=1, value=result.file_path)
            ws.cell(row=row, column=2, value=result.file_type)
            ws.cell(row=row, column=3, value=result.status.value)
            ws.cell(row=row, column=4, value="O" if result.applicable else "X")
            ws.cell(row=row, column=5, value=len(result.changes))
            ws.cell(row=row, column=6, value=len(result.issues))
            row += 1
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
    
    def _create_issues_sheet(self, results: List[FileMigrationResult]):
        """이슈 시트 생성"""
        ws = self.wb.create_sheet("이슈")
        
        # 헤더
        headers = ["파일 경로", "심각도", "메시지", "제안", "행 번호"]
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # 데이터
        row = 2
        for result in results:
            for issue in result.issues:
                ws.cell(row=row, column=1, value=result.file_path)
                ws.cell(row=row, column=2, value=issue.severity)
                ws.cell(row=row, column=3, value=issue.message)
                ws.cell(row=row, column=4, value=issue.suggestion or "")
                ws.cell(row=row, column=5, value=issue.line_number or "")
                row += 1
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
    
    def save(self, output_path: str):
        """Excel 파일 저장"""
        if self.wb:
            self.wb.save(output_path)
